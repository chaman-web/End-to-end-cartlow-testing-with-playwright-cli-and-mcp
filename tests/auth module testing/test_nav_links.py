import pytest
import openpyxl
from openpyxl.styles import PatternFill, Font
from playwright.sync_api import Page

BASE_URL = "https://www.cartlow.com/uae/en"
DOMAIN = "https://www.cartlow.com"


def build_full_url(href: str) -> str | None:
    if not href or href.startswith(("#", "javascript", "callto:", "mailto:", "tel:")):
        return None
    return href if href.startswith("http") else f"{DOMAIN}{href}"


def collect_all_links(page: Page) -> list[dict]:
    """Collect links with section, parent category, and link text."""
    links = {}

    # NAV — track parent category by grouping
    for el in page.locator("header a[href], nav a[href]").all():
        try:
            href = el.get_attribute("href") or ""
            text = el.inner_text().strip()
            full = build_full_url(href)
            if not full or full in links:
                continue
            # Try to find parent nav item
            parent = el.evaluate("""e => {
                let p = e.closest('li, [class*="group"]');
                if (p) {
                    let heading = p.querySelector('span, a, button');
                    return heading ? heading.innerText.trim() : '';
                }
                return '';
            }""") or ""
            links[full] = {"url": full, "text": text[:60], "section": "NAV", "parent": parent[:40]}
        except:
            pass

    # FOOTER
    for el in page.locator("footer a[href]").all():
        try:
            href = el.get_attribute("href") or ""
            text = el.inner_text().strip()
            full = build_full_url(href)
            if not full or full in links:
                continue
            # Try to find footer section heading
            parent = el.evaluate("""e => {
                let section = e.closest('div, section, ul');
                if (section) {
                    let h = section.querySelector('h2, h3, h4, p[class*="font"], span[class*="font"]');
                    return h ? h.innerText.trim() : '';
                }
                return '';
            }""") or ""
            links[full] = {"url": full, "text": text[:60], "section": "FOOTER", "parent": parent[:40]}
        except:
            pass

    # SIDEBAR
    for el in page.locator("aside a[href], [class*='sidebar'] a[href], [class*='category'] a[href]").all():
        try:
            href = el.get_attribute("href") or ""
            text = el.inner_text().strip()
            full = build_full_url(href)
            if not full or full in links:
                continue
            links[full] = {"url": full, "text": text[:60], "section": "SIDEBAR", "parent": ""}
        except:
            pass

    return list(links.values())


def check_status(url: str, page: Page) -> int:
    """Check URL status using the existing page's API request context."""
    try:
        r = page.request.get(url, timeout=15000)
        return r.status
    except Exception as e:
        print(f"  ⚠️  {url}: {e}")
        return 0


def save_excel(all_links: list[dict], broken: list[dict], ok: list[dict]):
    wb = openpyxl.Workbook()

    red_fill = PatternFill("solid", fgColor="FFCCCC")
    green_fill = PatternFill("solid", fgColor="CCFFCC")
    header_font = Font(bold=True)

    # Sheet 1: Broken links
    ws1 = wb.active
    ws1.title = "Broken Links"
    headers = ["Status Code", "URL", "Link Text", "Section", "Parent Category"]
    ws1.append(headers)
    for cell in ws1[1]:
        cell.font = header_font
    for row in broken:
        ws1.append([row["status"], row["url"], row["text"], row["section"], row["parent"]])
        for cell in ws1[ws1.max_row]:
            cell.fill = red_fill

    # Sheet 2: All links
    ws2 = wb.create_sheet("All Links")
    ws2.append(headers)
    for cell in ws2[1]:
        cell.font = header_font
    for row in all_links:
        ws2.append([row["status"], row["url"], row["text"], row["section"], row["parent"]])
        fill = green_fill if row["status"] and row["status"] < 400 else red_fill
        for cell in ws2[ws2.max_row]:
            cell.fill = fill

    # Auto-width columns
    for ws in [ws1, ws2]:
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 80)

    path = "reports/link_check_production.xlsx"
    wb.save(path)
    return path


def test_all_homepage_links(page: Page):
    """Check all nav/footer/sidebar links on production and save Excel report."""
    page.goto(BASE_URL, wait_until="domcontentloaded")
    page.wait_for_timeout(8000)

    # Close the app download popup if it appears
    try:
        close_btn = page.locator("button[aria-label='Close']").last
        if close_btn.is_visible():
            close_btn.click()
            page.wait_for_timeout(500)
    except:
        pass

    # Hover nav items to reveal dropdowns
    for nav_item in page.locator("header a, nav > ul > li").all():
        try:
            nav_item.hover()
            page.wait_for_timeout(200)
        except:
            pass

    links = collect_all_links(page)
    print(f"\n📋 Total unique links found: {len(links)}")

    all_results = []
    broken = []
    ok = []

    for link in links:
        status = check_status(link["url"], page)
        link["status"] = status
        all_results.append(link)
        if status == 0 or status >= 400:
            broken.append(link)
            print(f"  ❌ [{status}] {link['section']} | {link['parent']} | {link['text']} → {link['url']}")
        else:
            ok.append(link)
            print(f"  ✅ [{status}] {link['text']} → {link['url']}")

    report_path = save_excel(all_results, broken, ok)
    print(f"\n📊 Excel report saved to {report_path}")
    print(f"✅ OK: {len(ok)} | ❌ Broken: {len(broken)}")

    assert len(broken) == 0, f"{len(broken)} broken link(s) found. See {report_path}"
